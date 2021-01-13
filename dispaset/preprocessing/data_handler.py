import datetime as dt
import logging
import os
import sys

import numpy as np
import pandas as pd

from ..common import commons
try:
    from future.builtins import int
except ImportError:
    pass

DEFAULTS = {'ReservoirLevelInitial':0.5,'ReservoirLevelFinal':0.5,'ValueOfLostLoad':1E5,
                'PriceOfSpillage':1,'WaterValue':100,'ShareOfQuickStartUnits':0.5,
                'PriceOfNuclear':0,'PriceOfBlackCoal':0,'PriceOfGas':0,'PriceOfFuelOil':0,'PriceOfBiomass':0,
                'PriceOfCO2':0,'PriceOfLignite':0,'PriceOfPeat':0,'LoadShedding':0,'CostHeatSlack':0,
                'CostLoadShedding':100,'ShareOfFlexibleDemand':0,'DemandFlexibility':0,'PriceTransmission':0}

def NodeBasedTable(varname,config,default=None):
    '''
    This function loads the tabular data stored in csv files relative to each
    zone of the simulation.

    :param varname:             Variable name (as defined in config)
    :param idx:                 Pandas datetime index to be used for the output
    :param zones:               List with the zone codes to be considered
    :param fallback:            List with the order of data source.
    :param default:             Default value to be applied if no data is found

    :return:           Dataframe with the time series for each unit
    '''
    
    path = config[varname]
    zones=config['zones']       
    paths = {}
    if os.path.isfile(path):
        paths['all'] = path
        SingleFile=True
    elif '##' in path:
        for z in zones:
            path_c = path.replace('##', str(z))
            if os.path.isfile(path_c):
                paths[str(z)] = path_c
            else:
                logging.critical('No data file found for the table ' + varname + ' and zone ' + z + '. File ' + path_c + ' does not exist')
                sys.exit(1)
        SingleFile=False
    elif path != '':
        logging.critical('A path has been specified for table ' + varname + ' (' + path + ') but no file has been found')
        sys.exit(1)
    data = pd.DataFrame(index=config['idx_long'])
    if len(paths) == 0:
        logging.info('No data file specified for the table ' + varname + '. Using default value ' + str(default))
        if default is None:
            pass
        elif isinstance(default,(float,int)):
            data = pd.DataFrame(default,index=config['idx_long'],columns=zones)
        else:
            logging.critical('Default value provided for table ' + varname + ' is not valid')
            sys.exit(1)
    elif SingleFile:
        # If it is only one file, there is a header with the zone code
        tmp = load_time_series(config,paths['all'])
           
        if len(tmp.columns) == 1:    # if there is only one column, assign its value to all the zones, whatever the header
            try:    # if the column header is numerical, there was probably no header. Load the file again.
                float(tmp.columns[0])   # this will fail if the header is not numerical
                tmp = load_time_series(config,paths['all'],header=None)
            except:
                pass
            for key in zones:
                data[key] = tmp.iloc[:,0]
        else:
            for key in zones:
                if key in tmp:
                    data[key] = tmp[key]
                else:
                    logging.error('Zone ' + key + ' could not be found in the file ' + path + '. Using default value ' + str(default))
                    if default is None:
                        pass
                    elif isinstance(default,(float,int)):
                        data[key] = default
                    else:
                        logging.critical('Default value provided for table ' + varname + ' is not valid')
                        sys.exit(1)
    else: # assembling the files in a single dataframe:
        for z in paths:
            # In case of separated files for each zone, there is no header
            tmp = load_time_series(config,paths[z])  
            data[z] = tmp.iloc[:,0]

    return data


def UnitBasedTable(plants,varname,config,fallbacks=['Unit'],default=None,RestrictWarning=None):
    '''
    This function loads the tabular data stored in csv files and assigns the
    proper values to each unit of the plants dataframe. If the unit-specific 
    value is not found in the data, the script can fallback on more generic
    data (e.g. fuel-based, technology-based, zone-based) or to the default value.
    The order in which the data should be loaded is specified in the fallback
    list. For example, ['Unit','Technology'] means that the script will first
    try to find a perfect match for the unit name in the data table. If not found,
    a column with the unit technology as header is search. If not found, the
    default value is assigned.

    :param plants:              Dataframe with the units for which data is required
    :param varname:             Variable name (as defined in config)
    :param idx:                 Pandas datetime index to be used for the output
    :param zones:           List with the zone codes to be considered
    :param fallback:            List with the order of data source. 
    :param default:             Default value to be applied if no data is found
    :param RestrictWarning:     Only display the warnings if the unit belongs to the list of technologies provided in this parameter
    
    :return:           Dataframe with the time series for each unit
    '''
    path = config[varname]   
    zones = config['zones']    
    paths = {}
    if os.path.isfile(path):
        paths['all'] = path
        SingleFile=True
    elif '##' in path:
        for z in zones:
            path_c = path.replace('##', str(z))
            if os.path.isfile(path_c):
                paths[str(z)] = path_c
            else:
                logging.error('No data file found for the table ' + varname + ' and zone ' + z + '. File ' + path_c + ' does not exist')
#                sys.exit(1)
        SingleFile=False
    elif path != '':
        logging.critical('A path has been specified for table ' + varname + ' (' + path + ') but no file has been found')
        sys.exit(1)

    data = pd.DataFrame(index=config['idx_long'])
    if len(paths) == 0:
        logging.info('No data file specified for the table ' + varname + '. Using default value ' + str(default))
        if default is None:
            out = pd.DataFrame(index=config['idx_long'])
        elif isinstance(default,(float,int)):
            out = pd.DataFrame(default,index=config['idx_long'],columns=plants['Unit'])
        else:
            logging.critical('Default value provided for table ' + varname + ' is not valid')
            sys.exit(1)
    else: # assembling the files in a single dataframe:
        columns = []
        for z in paths:
            tmp = load_time_series(config,paths[z])
            if SingleFile:
                for key in tmp:
                    data[key] = tmp[key]                
            else:    # use the multi-index header with the zone
                for key in tmp:
                    columns.append((z,key))
                    data[z+','+key] = tmp[key]
        if not SingleFile:
            data.columns = pd.MultiIndex.from_tuples(columns, names=['Zone', 'Data'])
        # For each plant and each fallback key, try to find the corresponding column in the data
        out = pd.DataFrame(index=config['idx_long'])
        for j in plants.index:
            warning = True
            if not RestrictWarning is None:
                warning = False
                if plants.loc[j,'Technology'] in RestrictWarning:
                    warning=True
            u = plants.loc[j,'Unit']
            found = False
            for i,key in enumerate(fallbacks):
                if SingleFile:
                    header = plants.loc[j,key]
                else:
                    header = (plants.loc[j,'Zone'],plants.loc[j,key])
                if header in data:
                    out[u] = data[header]
                    found = True
                    if i > 0 and warning:
                        logging.warning('No specific information was found for unit ' + u + ' in table ' + varname + '. The generic information for ' + str(header) + ' has been used')
                    break
            if not found:
                if warning:
                    logging.info('No specific information was found for unit ' + u + ' in table ' + varname + '. Using default value ' + str(default))
                if not default is None:
                    out[u] = default
    if not out.columns.is_unique:
        logging.critical('The column headers of table "' + varname + '" are not unique!. The following headers are duplicated: ' + str(out.columns.get_duplicates()))
        sys.exit(1)
    return out


def merge_series(plants,oldplants, data, method='WeightedAverage', tablename=''):
    """
    Function that merges the times series corresponding to the merged units (e.g. outages, inflows, etc.)

    :param plants:      Pandas dataframe with final units after clustering (must contain 'FormerUnits')
    :param oldplants:   Pandas dataframe with the original units
    :param data:        Pandas dataframe with the time series and the original unit names as column header
    :param method:      Select the merging method ('WeightedAverage'/'Sum')
    :param tablename:   Name of the table being processed (e.g. 'Outages'), used in the warnings
    :return merged:     Pandas dataframe with the merged time series when necessary
    """
    # backward compatibility:
    if not "Nunits" in plants:
        plants['Nunits'] = 1
        
    if not 'FormerUnits' in plants:
        logging.critical('The unit table provided must contain the columns "FormerUnits"')
        sys.exit(1)

    merged = pd.DataFrame(index=data.index)
    
    # Create a dictionary relating the former units to the new (clustered) ones:
    units = {}
    for u in plants.index:
        for uu in plants.loc[u,'FormerUnits']:
            units[uu] = u
        
    # First check the data:
    if not isinstance(data,pd.DataFrame):
        logging.critical('The input "' + tablename + '" to the merge_series function must be a dataframe')
        sys.exit(1)
    for key in data:
        if str(data[key].dtype) not in ['bool','int','float','float16', 'float32', 'float64', 'float128','int8', 'int16', 'int32', 'int64']:
            logging.critical('The column "' + str(key) + '" of table + "' + tablename + '" is not numeric!')
    for key in data:
        if key in units:
            newunit = units[key]
            if newunit not in merged:  # if the columns name is in the mapping and the new unit has not been processed yet
                oldnames = plants.loc[newunit,'FormerUnits']
                if all([name in data for name in oldnames]):
                    subunits = data[oldnames]
                else:
                    for name in oldnames:
                        if name not in data:
                            logging.critical('The column "' + name + '" is required for the aggregation of unit "' + key +
                                             '", but it has not been found in the input data')
                            sys.exit(1)
                value = np.zeros(len(data))
                # Renaming the subunits df headers with the old plant indexes instead of the unit names:
                if method == 'WeightedAverage':
                    for name in oldnames:
                        value = value + subunits[name] * np.maximum(1e-9, oldplants['PowerCapacity'][name]*oldplants['Nunits'][name])
                    P_j = np.sum(np.maximum(1e-9, oldplants['PowerCapacity'][oldnames]*oldplants['Nunits'][oldnames]))
                    merged[newunit] = value / P_j
                elif method == 'StorageWeightedAverage':
                    for name in oldnames:
                       value = value + subunits[name] * np.maximum(1e-9, oldplants['STOCapacity'][name]*oldplants['Nunits'][name])
                    P_j = np.sum(np.maximum(1e-9, oldplants['STOCapacity'][oldnames]*oldplants['Nunits'][oldnames]))
                    merged[newunit] = value / P_j
                elif method == 'Sum':
                    merged[newunit] = subunits.sum(axis=1)
                else:
                    logging.critical('Method "' + str(method) + '" unknown in function MergeSeries')
                    sys.exit(1)
        elif key in oldplants['Unit']:
            if not isinstance(key, tuple):  # if the columns header is a tuple, it does not come from the data and has been added by Dispa-SET
                logging.warning('Column ' + str(key) + ' present in the table "' + tablename + '" not found in the mapping between original and clustered units. Skipping')
        else:
            if not isinstance(key, tuple):  # if the columns header is a tuple, it does not come from the data and has been added by Dispa-SET
                logging.warning('Column ' + str(key) + ' present in the table "' + tablename + '" not found in the table of power plants. Skipping')
    return merged


def define_parameter(sets_in, sets, value=0):
    """
    Function to define a DispaSET parameter and fill it with a constant value

    :param sets_in:     List with the labels of the sets corresponding to the parameter
    :param sets:        dictionary containing the definition of all the sets (must comprise those referenced in sets_in)
    :param value:       Default value to attribute to the parameter
    """
    if value == 'bool':
        values = np.zeros([len(sets[setx]) for setx in sets_in], dtype='bool')
    elif value == 0:
        values = np.zeros([len(sets[setx]) for setx in sets_in])
    elif value == 1:
        values = np.ones([len(sets[setx]) for setx in sets_in])
    else:
        values = np.ones([len(sets[setx]) for setx in sets_in]) * value
    return {'sets': sets_in, 'val': values}


def load_time_series(config,path,header='infer'):
    """
    Function that loads time series data, checks the compatibility of the indexes
    and guesses when no exact match between the required index and the data is 
    present
    """

    data = pd.read_csv(path, index_col=0, parse_dates=True, header=header)
    
    if not data.index.is_unique:
        logging.critical('The index of data file ' + path + ' is not unique. Please check the data')
        sys.exit(1)

    if not data.index.is_monotonic_increasing:
        logging.error('The index of data file ' + path + ' is not monotoneously increasing. Trying to check if it can be parsed with a "day first" format ')
        data = pd.read_csv(path, index_col=0, parse_dates=True, header=header, dayfirst=True)
        if not data.index.is_monotonic_increasing:
            logging.critical('Could not parse index of ' + path + '. To avoid problems make sure that you use the proper american date format (yyyy-mm-dd hh:mm:ss)')
            sys.exit(1)
        
    # First convert numerical indexes into datetimeindex:
    if data.index.is_numeric():
        if len(data) == len(config['idx']):  # The data has the same length as the provided index range
            logging.info('A numerical index has been found for file ' + path + 
                         '. It has the same length as the target simulation and is therefore considered valid')
            data.index=config['idx']
        elif len(data) == 8760:
            logging.info('A numerical index has been found for file ' + path + 
                         '. Since it contains 8760 elements, it is assumed that it corresponds to a whole year')
            data.index = pd.date_range(start=dt.datetime(*(config['idx'][0].year,1,1,0,0)),
                                                        end=dt.datetime(*(config['idx'][0].year,12,31,23,59,59)),
                                                        freq=commons['TimeStep'])
        else:
            logging.critical('A numerical index has been found for file ' + path + 
                         '. However, its length does not allow guessing its timestamps. Please use a 8760 elements time series')
            sys.exit(1)

    if data.index.is_all_dates:
        data.index = data.index.tz_localize(None)   # removing locational data
        # Checking if the required index entries are in the data:
        common = data.index.intersection(config['idx'])
        if len(common) == 0:
            # check if original year is leap year and destination year is not (remove leap date)
            if (data.index[0].is_leap_year is True) and (config['idx'][0].is_leap_year is False):
                data = data[~((data.index.month == 2) & (data.index.day == 29))]
                logging.warning('File ' + path + ': data for year ' + str(data.index[0].year) +
                                ' is used instead of year ' + str(config['idx'][0].year))
                data.index = data.index.map(lambda t: t.replace(year=config['idx'][0].year))
            # check if both years are either leap or non leap
            elif (data.index[0].is_leap_year is True) and (config['idx'][0].is_leap_year is True) or \
                    (data.index[0].is_leap_year is False) and (config['idx'][0].is_leap_year is False):
                logging.warning('File ' + path + ': data for year ' + str(data.index[0].year) +
                                ' is used instead of year ' + str(config['idx'][0].year) +
                                '. Leap year date is removed from the original DataFrame.')
                data.index = data.index.map(lambda t: t.replace(year=config['idx'][0].year))
            # check if original year is not a leap year and destination year is a leap year (add leap date and take average hourly values between 28.02. and 1.3.
            elif (data.index[0].is_leap_year is False) and (config['idx'][0].is_leap_year is True):
                logging.warning('File ' + path + ': data for year ' + str(data.index[0].year) +
                                ' is used instead of year ' + str(config['idx'][0].year) +
                                '. Leap year date is interpolated between the two neighbouring days.')
                data.index = data.index.map(lambda t: t.replace(year=config['idx'][0].year))
                mask = data.loc[str(config['idx'][0].year)+'-2-28': str(config['idx'][0].year)+'-3-1']
                mask = mask.groupby(mask.index.hour).mean()
                time = pd.date_range(str(config['idx'][0].year)+'-2-29', periods=24, freq='H')
                mask = mask.set_index(time)
                data = data.reindex(config['idx'])
                data.update(mask)
        # recompute common index entries, and check again:
        common = data.index.intersection(config['idx'])
        if len(common) < len(config['idx'])-1:
            logging.critical('File ' + path + ': the index does not contain the necessary time range (from ' + str(config['idx'][0]) + ' to ' + str(config['idx'][-1]) + ')')
            sys.exit(1)
        elif len(common) == len(config['idx'])-1:  # there is only one data point missing. This is deemed acceptable
            logging.warning('File ' + path + ': there is one data point missing in the time series. It will be filled with the nearest data')
        else:
            pass              # the defined simulation index is found within the data. No action required
    else:
        logging.critical('Index for file ' + path + ' is not valid')
        sys.exit(1)
        
    # re-indexing with the longer index (including look-ahead) and filling possibly missing data at the beginning and at the end::
    return data.reindex(config['idx_long'], method='nearest').fillna(method='bfill')


def load_config(ConfigFile,AbsPath=True):
    """
    Wrapper function around load_config_excel and load_config_yaml
    """
    if ConfigFile.endswith(('.xlsx','.xls')):
        config = load_config_excel(ConfigFile,AbsPath=True)
    elif ConfigFile.endswith(('.yml','.yaml')):
        config = load_config_yaml(ConfigFile,AbsPath=True)
    else:
        logging.critical('The extension of the config file should be .xlsx or .yml')
        sys.exit(1)
    return config

def read_truefalse(sheet, rowstart, colstart, rowstop, colstop, colapart=1):
    """
    Function that reads a two column format with a list of strings in the first
    columns and a list of true false in the second column
    The list of strings associated with a True value is returned
    """
    out = []
    for i in range(rowstart, rowstop+1):
        if sheet.cell(i, colstart + colapart).internal_value == 1:
            out.append(sheet.cell(i, colstart).internal_value)
    return out

def load_config_excel(ConfigFile,AbsPath=True):
    """
    Function that loads the DispaSET excel config file and returns a dictionary
    with the values

    :param ConfigFile: String with (relative) path to the DispaSET excel configuration file
    :param AbsPath:    If true, relative paths are automatically changed into absolute paths (recommended)
    """
    from openpyxl import load_workbook
    
    wb = load_workbook(ConfigFile)  # Option for csv to be added later
    sheet = wb['main']
    config = {}
    
    if sheet.cell(1,1).internal_value == 'Dispa-SET Configuration File (v20.01)':
        config['Description'] = sheet.cell(6,2).internal_value
        config['StartDate'] = sheet.cell(57,3).internal_value.timetuple()[0:6]
        config['StopDate'] = sheet.cell(58,3).internal_value.timetuple()[0:6]
        config['HorizonLength'] = int(sheet.cell(59,3).internal_value)
        config['LookAhead'] = int(sheet.cell(60,3).internal_value)
        
        # Defning the input locations in the config file:
        StdParameters={'SimulationDirectory':33,'WriteGDX':34,'WritePickle':35,'GAMS_folder':36,
                          'cplex_path':37,'DataTimeStep':60,'SimulationTimeStep':61,
                          'SimulationType':76,'ReserveCalculation':77,'AllowCurtailment':78,
                          'HydroScheduling':98,'HydroSchedulingHorizon':99,'InitialFinalReservoirLevel':100}
        PathParameters={'Demand':124, 'Outages':126, 'PowerPlantData':127, 'RenewablesAF':128, 
                          'LoadShedding':129, 'NTC':130, 'Interconnections':131, 'ReservoirScaledInflows':132, 
                          'PriceOfNuclear':180, 'PriceOfBlackCoal':181, 'PriceOfGas':182, 
                          'PriceOfFuelOil':183,'PriceOfBiomass':184, 'PriceOfCO2':166, 
                          'ReservoirLevels':133, 'PriceOfLignite':185, 'PriceOfPeat':186,
                          'HeatDemand':134,'CostHeatSlack':165,'CostLoadShedding':168,'ShareOfFlexibleDemand':125,
                          'Temperatures':135,'PriceTransmission':169,'Reserve2U':160,'Reserve2D':161}
        modifiers= {'Demand':274,'Wind':275,'Solar':276,'Storage':277}
        default = {'ReservoirLevelInitial':101,'ReservoirLevelFinal':102,'PriceOfNuclear':180,'PriceOfBlackCoal':181,
                    'PriceOfGas':182,'PriceOfFuelOil':183,'PriceOfBiomass':184,'PriceOfCO2':166,'PriceOfLignite':185,
                    'PriceOfPeat':186,'LoadShedding':129,'CostHeatSlack':167,'CostLoadShedding':168,'ValueOfLostLoad':204,
                    'PriceOfSpillage':205,'WaterValue':206,'ShareOfQuickStartUnits':163,'ShareOfFlexibleDemand':125,
                    'DemandFlexibility':162,'PriceTransmission':169}
        for p in StdParameters:
            config[p] = sheet.cell(StdParameters[p] + 1, 3).internal_value
        for p in PathParameters:
            config[p] = sheet.cell(PathParameters[p] + 1, 3).internal_value
        config['modifiers'] = {}
        for p in modifiers:
            config['modifiers'][p] = sheet.cell(modifiers[p] + 1, 3).internal_value
        config['default'] = {}
        for p in default:
            config['default'][p] = sheet.cell(default[p] + 1, 6).internal_value
            
        #True/Falst values:
        config['zones'] = read_truefalse(sheet, 226, 2, 247, 4)
        config['zones'] = config['zones'] + read_truefalse(sheet, 226, 5, 247, 7)
        config['mts_zones'] = read_truefalse(sheet, 226, 2, 247, 4, 2)
        config['mts_zones'] = config['mts_zones'] + read_truefalse(sheet, 226, 5, 247, 7, 2)
        config['ReserveParticipation'] = read_truefalse(sheet, 306, 2, 319, 4)

        # Set default values (for backward compatibility):
        for param in DEFAULTS:
            if config['default'][param]=='':
                config['default'][param]=DEFAULTS[param]
                logging.warning('No value was provided in config file for {}. Will use {}'.format(param, DEFAULTS[param]))
                config['default'][param] = DEFAULTS[param]

        if AbsPath:
        # Changing all relative paths to absolute paths. Relative paths must be defined 
        # relative to the parent folder of the config file.
            abspath = os.path.abspath(ConfigFile)
            basefolder = os.path.abspath(os.path.join(os.path.dirname(abspath),os.pardir))
            if not os.path.isabs(config['SimulationDirectory']):
                config['SimulationDirectory'] = os.path.join(basefolder,config['SimulationDirectory'])
            for param in PathParameters:
                if config[param] == '' or config[param] is None or str(config[param]).isspace():
                    config[param] = ''
                elif not os.path.isabs(config[param]):
                    config[param] = os.path.join(basefolder,config[param])

        logging.info("Using config file (v20.01) " + ConfigFile + " to build the simulation environment")
        logging.info("Using " + config['SimulationDirectory'] + " as simulation folder")
        logging.info("Description of the simulation: "+ config['Description'])
        
        return config        
        

    elif sheet.cell(1,1).internal_value == 'Dispa-SET Configuration File':
        config['Description'] = sheet.cell(6, 2).internal_value
        config['SimulationDirectory'] = sheet.cell(18, 3).internal_value
        config['WriteExcel'] = sheet.cell(19, 3).internal_value
        config['WriteGDX'] = sheet.cell(20, 3).internal_value
        config['WritePickle'] = sheet.cell(21, 3).internal_value
        config['GAMS_folder'] = sheet.cell(22, 3).internal_value
        config['cplex_path'] = sheet.cell(23, 3).internal_value
    
        config['StartDate'] = sheet.cell(31, 3).internal_value.timetuple()[0:6]
        config['StopDate'] = sheet.cell(32, 3).internal_value.timetuple()[0:6]
        config['HorizonLength'] = int(sheet.cell(33, 3).internal_value)
        config['LookAhead'] = int(sheet.cell(34, 3).internal_value)
        config['DataTimeStep'] = sheet.cell(35, 3).internal_value
        config['SimulationTimeStep'] = sheet.cell(36, 3).internal_value
    
        config['SimulationType'] = sheet.cell(47, 3).internal_value
        config['ReserveCalculation'] = sheet.cell(48, 3).internal_value
        config['AllowCurtailment'] = sheet.cell(49, 3).internal_value
    
        config['HydroScheduling'] = sheet.cell(54, 3).internal_value
        config['HydroSchedulingHorizon'] = sheet.cell(55, 3).internal_value
        config['InitialFinalReservoirLevel'] = sheet.cell(56, 3).internal_value
    
        # Set default values (for backward compatibility):
        NonEmptyarameters = {'DataTimeStep':1,'SimulationTimeStep':1,'HydroScheduling':'Off','HydroSchedulingHorizon':'Annual','InitialFinalReservoirLevel':True}
        for param in NonEmptyarameters:
            if config[param]=='':
                config[param]=NonEmptyarameters[param]   
    
        # List of parameters for which an external file path must be specified:
        PARAMS = ['Demand', 'Outages', 'PowerPlantData', 'RenewablesAF', 'LoadShedding', 'NTC', 'Interconnections',
          'ReservoirScaledInflows', 'PriceOfNuclear', 'PriceOfBlackCoal', 'PriceOfGas', 'PriceOfFuelOil',
          'PriceOfBiomass', 'PriceOfCO2', 'ReservoirLevels', 'PriceOfLignite', 'PriceOfPeat','HeatDemand',
          'CostHeatSlack','CostLoadShedding','ShareOfFlexibleDemand']
        for i, param in enumerate(PARAMS):
            config[param] = sheet.cell(62 + i, 3).internal_value
    
        # List of new parameters for which an external file path must be specified:
        params2 = ['Temperatures','PriceTransmission','Reserve2D','Reserve2U']
        if sheet.max_row>150:                 # for backward compatibility (old excel sheets had less than 150 rows)
            for i, param in enumerate(params2):
                config[param] = sheet.cell(157 + i, 3).internal_value
        else:
            for param in params2:
                config[param] = ''
    
        if AbsPath:
        # Changing all relative paths to absolute paths. Relative paths must be defined 
        # relative to the parent folder of the config file.
            abspath = os.path.abspath(ConfigFile)
            basefolder = os.path.abspath(os.path.join(os.path.dirname(abspath),os.pardir))
            if not os.path.isabs(config['SimulationDirectory']):
                config['SimulationDirectory'] = os.path.join(basefolder,config['SimulationDirectory'])
            for param in PARAMS+params2:
                if config[param] == '' or config[param] is None or str(config[param]).isspace():
                    config[param] = ''
                elif not os.path.isabs(config[param]):
                    config[param] = os.path.join(basefolder,config[param])
    
        config['default'] = {}
        config['default']['ReservoirLevelInitial'] = sheet.cell(57, 6).internal_value
        config['default']['ReservoirLevelFinal'] = sheet.cell(58, 6).internal_value
        config['default']['PriceOfNuclear'] = sheet.cell(70, 6).internal_value
        config['default']['PriceOfBlackCoal'] = sheet.cell(71, 6).internal_value
        config['default']['PriceOfGas'] = sheet.cell(72, 6).internal_value
        config['default']['PriceOfFuelOil'] = sheet.cell(73, 6).internal_value
        config['default']['PriceOfBiomass'] = sheet.cell(74, 6).internal_value
        config['default']['PriceOfCO2'] = sheet.cell(75, 6).internal_value
        config['default']['PriceOfLignite'] = sheet.cell(77, 6).internal_value
        config['default']['PriceOfPeat'] = sheet.cell(78, 6).internal_value
        config['default']['LoadShedding'] = sheet.cell(66, 6).internal_value
        config['default']['CostHeatSlack'] = sheet.cell(80, 6).internal_value
        config['default']['CostLoadShedding'] = sheet.cell(81, 6).internal_value
        config['default']['ValueOfLostLoad'] = sheet.cell(82, 6).internal_value
        config['default']['PriceOfSpillage'] = sheet.cell(83, 6).internal_value
        config['default']['WaterValue'] = sheet.cell(84, 6).internal_value
        config['default']['ShareOfQuickStartUnits'] = 0.5          # to be added to xlsx file
        
        # Set default values (for backward compatibility):
        for param in DEFAULTS:
            if config['default'].get(param,'')=='':
                config['default'][param]=DEFAULTS[param]
    
        config['zones'] = read_truefalse(sheet, 87, 2, 110, 4)
        config['zones'] = config['zones'] + read_truefalse(sheet, 87, 5, 110, 7)
    
        config['mts_zones'] = read_truefalse(sheet, 87, 2, 110, 4, 2)
        config['mts_zones'] = config['mts_zones'] + read_truefalse(sheet, 87, 5, 110, 7, 2)
    
        config['modifiers'] = {}
        config['modifiers']['Demand'] = sheet.cell(112, 3).internal_value
        config['modifiers']['Wind'] = sheet.cell(113, 3).internal_value
        config['modifiers']['Solar'] = sheet.cell(114, 3).internal_value
        config['modifiers']['Storage'] = sheet.cell(115, 3).internal_value
    
        # Read the technologies participating to reserve markets:
        config['ReserveParticipation'] = read_truefalse(sheet, 132, 2, 146, 4)
    
        logging.info("Using config file " + ConfigFile + " to build the simulation environment")
        logging.info("Using " + config['SimulationDirectory'] + " as simulation folder")
        logging.info("Description of the simulation: "+ config['Description'])
        
        return config
    
    else:
        logging.critical('The format of the excel config file (defined by its main title) is not recognized')
        sys.exit(1)

def load_config_yaml(filename, AbsPath=True):
    """ Loads YAML file to dictionary"""
    import yaml
    with open(filename, 'r') as f:
        try:
            config = yaml.full_load(f)
        except yaml.YAMLError as exc:
            logging.error('Cannot parse config file: {}'.format(filename))
            raise exc
            
    # List of parameters to be added with a default value if not present (for backward compatibility):
    
    params_to_be_added = {'Temperatures':'','DataTimeStep':1,'SimulationTimeStep':1,'HydroScheduling':'Off','HydroSchedulingHorizon':'Annual','InitialFinalReservoirLevel':True}
    for param in params_to_be_added:
        if param not in config:
            config[param] = params_to_be_added[param]
                        
    # Set default values (for backward compatibility):
    NonEmptyDefaultss = {'ReservoirLevelInitial':0.5,'ReservoirLevelFinal':0.5,'ValueOfLostLoad':1E5,'PriceOfSpillage':1,'WaterValue':100,'ShareOfQuickStartUnits':0.5}
    for param in NonEmptyDefaultss:
        if param not in config['default']:
            config['default'][param]=NonEmptyDefaultss[param]


    # Define missing parameters if they were not provided in the config file
    PARAMS = ['Demand', 'Outages', 'PowerPlantData', 'RenewablesAF', 'LoadShedding', 'NTC', 'Interconnections',
          'ReservoirScaledInflows', 'PriceOfNuclear', 'PriceOfBlackCoal', 'PriceOfGas', 'PriceOfFuelOil',
          'PriceOfBiomass', 'PriceOfCO2', 'ReservoirLevels', 'PriceOfLignite', 'PriceOfPeat','HeatDemand',
          'CostHeatSlack','CostLoadShedding','ShareOfFlexibleDemand','Temperatures','PriceTransmission',
          'Reserve2D','Reserve2U']
    for param in PARAMS:
        if param not in config:
            config[param] = ''    
    global DEFAULTS
    for key in DEFAULTS:
        if key not in config['default']:
            config['default'][key]=DEFAULTS[key]

    if AbsPath:
    # Changing all relative paths to absolute paths. Relative paths must be defined 
    # relative to the parent folder of the config file.
        abspath = os.path.abspath(filename)
        basefolder = os.path.abspath(os.path.join(os.path.dirname(abspath),os.pardir))
        if not os.path.isabs(config['SimulationDirectory']):
            config['SimulationDirectory'] = os.path.join(basefolder,config['SimulationDirectory'])
        for param in PARAMS:
            if not os.path.isabs(config[param]):
                if config[param] == '' or config[param].isspace():
                    config[param] = ''
                elif not os.path.isabs(config[param]):
                    config[param] = os.path.join(basefolder,config[param])
    return config

def export_yaml_config(ExcelFile, YAMLFile):
    """
    Function that loads the DispaSET excel config file and dumps it as a yaml file.

    :param ExcelFile:   Path to the Excel config file
    :param YAMLFile:    Path to the YAML config file to be written
    """
    import yaml
    config = load_config_excel(ExcelFile,AbsPath=False)
    with open(YAMLFile, 'w') as outfile:
        yaml.dump(config, outfile, default_flow_style=False)
    return True


